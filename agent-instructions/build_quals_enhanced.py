"""Enhance the extracted QUALS MASTER with capability / service-area mapping.

Reads the flat extraction (Slide, Title, Category, Client Overview, Business Situation,
Engagement, Technologies, Outcomes, Notes) and adds the columns a pursuit agent needs to
answer "give me the strongest credentials for <industry> + <service area>".

Outcomes are merged only where an engagement was matched by hand against
Aberdeen_Healthcare_Tech_Quals. Nothing is invented; unmatched rows are flagged for owner
input rather than filled.
"""
import collections
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else "github_quals.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "AberdeenAdv_QUALS_MASTER_enhanced_081426.xlsx"

n = lambda v: "" if v is None else str(v).replace("\n", " ").strip()

rows = [tuple(n(c) for c in r) + ("",) * 9
        for r in openpyxl.load_workbook(SRC, data_only=True)["Quals Master"]
        .iter_rows(min_row=2, values_only=True)]
rows = [r for r in rows if r[1]]

# --- derivation: keyword -> (capability, service area) --------------------
# Ordered; first match wins. Deliberately conservative - an unmatched row is
# flagged, not guessed.
CAP = [
    (r"due diligence|post-merger|m&a|merger|acquisition|target overview",
     "Deal Enablement & Growth", "M&A Enablement"),
    (r"cost benchmark|expense management|travel and expense|cogs|pricing",
     "Financial Transparency", "IT Financial Mgmt. & Optimization"),
    (r"epic|erp|s/4|implementation|migration|consolidation|cutover|quality assurance"
     r"|transformation office|tmo|pmo|program",
     "Platform Implementation & Modernization", "Large Program Assurance, Delivery & Change"),
    (r"portal|dashboard|test automation|qa software|automation \(rpa\)|blue prism|raid log"
     r"|registration|metrics portal|software testing|product development",
     "Product & Application Services", "Technology Strategy & Operations"),
    (r"rationalization|itsm|it service management|demand management|insourcing|captive"
     r"|process mapping|org design|organization design|operating model|resource planning"
     r"|field services",
     "Organizational & Operational Efficiency", "Technology Strategy & Operations"),
    (r"\bdata\b|analytics|powerbi|power bi|\bmdm\b|governance|reporting",
     "Strategy & Transformation", "Technology Future-Proofing"),
    (r"strategy|persona|voice of the customer|journey|agile|vision|comparative analysis"
     r"|location strategy|partnership",
     "Strategy & Transformation", "Technology Strategy & Operations"),
]

# CRM controlled vocabulary (AGENT_WORKFLOW.md section 1)
IND = [
    (r"healthcare|hospital|epic|clinical|oncology|pain|bioimaging|behavioral|provider"
     r"|health system|patient|medical|pharma|dermatolog", "Healthcare & Life Sciences"),
    (r"financial services|banking|bank app|fintech|money transfer|remittance|wealth",
     "Financial Services"),
    (r"retail|qsr|consumer", "Consumer & Retail"),
    (r"cloud service provider|digital marketing", "Technology & Software"),
    (r"transportation", "Transportation & Logistics"),
    (r"spice|food", "Agriculture & Food"),
    (r"engineering & construction|construction|equipment rental", "Real Estate & Construction"),
    (r"broadband|distributor|manufactur", "Manufacturing & Industrial"),
]


def derive(rules, text, width):
    t = text.lower()
    for rule in rules:
        if re.search(rule[0], t):
            return list(rule[1:])
    return [""] * width


def clean_cat(c):
    """Normalise the 26 dirty category values into a usable practice area."""
    c = re.sub(r"\s+", " ", c).strip()
    c = re.sub(r"\s*(Client & Target Overview|IN PROGRESS)\s*", "", c, flags=re.I).strip()
    if c in ("Client Sample 2", "7.0 Vendor Qualifications", ""):
        return ""
    c = re.split(r"\s{2,}", c)[0].strip()          # strip trailing technology lists
    c = c.replace("–", "-")
    if c.lower().startswith("advisory"):
        tail = re.sub(r"^advisory\s*[-—]?\s*", "", c, flags=re.I).strip()
        return "Advisory — " + tail if tail else "Advisory"
    return c


# Outcomes verified by hand against Aberdeen_Healthcare_Tech_Quals_081426.xlsx.
# Only high-confidence matches. Fuzzy matching produced false positives and was not trusted.
VERIFIED = {
    "41": ("18 modules implemented over 12 months; 9 stars achieved", "Payvider", "$1B - $5B", "STRONG"),
    "45": ("Objective of one patient / one record delivered post-merger", "Provider", "$5B+", "STRONG"),
    "46": ("Objective of one patient / one record delivered post-merger", "Provider", "$5B+", "STRONG"),
    "47": ("$50m+ savings over 5 years", "Provider", "$500M - $999M", "STRONG"),
    "5":  ("Delivery consistency across all initiatives; risks identified with recommendations", "Provider", "$5B+", "STRONG"),
    "6":  ("TMO stood up; business interruption minimised", "Provider", "$5B+", "STRONG"),
}
FLAG_DUP = {"46": "duplicate of slide 45", "49": "duplicate of slide 20",
            "52": "duplicate of slide 29", "51": "near-duplicate of slide 50"}
NOT_QUAL = {"27": "section divider, not a qual", "56": "contact slide, not a qual"}

HDR = ["Slide", "Title", "Practice Area (normalised)", "Capability", "Service Area", "Industry",
       "Client Type", "Client Revenue Band", "Quantified Outcome", "Proof Strength", "Status",
       "External Use", "Derivation", "Record Quality", "Client Overview", "Business Situation",
       "Our Engagement / Solution", "Technologies", "Additional Notes",
       "Category (raw, as extracted)"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Quals Master"
ws.append(HDR)

stats = collections.Counter()
cap_cov, ind_cov, cap_out = collections.Counter(), collections.Counter(), collections.Counter()

for r in rows:
    slide, title, rawcat = r[0], r[1], r[2]
    blob = " ".join([title, rawcat, r[3], r[4], r[5]])
    cap, sa = derive(CAP, blob, 2)
    ind = derive(IND, blob, 1)[0]
    v = VERIFIED.get(slide)
    outcome, ctype, rev, strength = v if v else ("", "", "", "")

    quality = []
    if slide in NOT_QUAL:
        quality.append(NOT_QUAL[slide])
    if slide in FLAG_DUP:
        quality.append(FLAG_DUP[slide])
    if not rawcat:
        quality.append("category missing in source")
    if not outcome and slide not in NOT_QUAL:
        quality.append("NO OUTCOME - owner input needed")
    if not cap and slide not in NOT_QUAL:
        quality.append("capability not derivable from title")

    status = ("n/a" if slide in NOT_QUAL
              else "In Progress" if "IN PROGRESS" in rawcat.upper() else "Delivered")
    ws.append([slide, title, clean_cat(rawcat), cap, sa, ind, ctype, rev, outcome,
               strength or ("" if slide in NOT_QUAL else "THIN"), status, "TBD",
               "verified vs Healthcare Tech Quals" if v else "derived from title/category",
               "; ".join(quality), r[3], r[4], r[5], r[6], r[8], rawcat])

    if slide not in NOT_QUAL:
        stats["quals"] += 1
        cap_cov[cap or "(not derivable)"] += 1
        ind_cov[ind or "(not derivable)"] += 1
        if outcome:
            stats["with_outcome"] += 1
            cap_out[cap or "(not derivable)"] += 1

# --- styling -------------------------------------------------------------
NAVY, AMBER, GREY = "FF09375F", "FFF7CE01", "FFF2F2F2"
for c in ws[1]:
    c.font = Font(name="Calibri", bold=True, color="FFFFFFFF", size=9)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 34
for i, w in enumerate([6, 46, 26, 32, 34, 26, 14, 16, 44, 13, 12, 11, 30, 38, 50, 50, 50, 26, 30, 30], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "C2"
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.font = Font(name="Calibri", size=9)
    rq = str(row[13].value or "")
    if "NO OUTCOME" in rq:
        row[8].fill = PatternFill("solid", fgColor=AMBER)
    if "duplicate" in rq or "not a qual" in rq:
        for c in row:
            c.fill = PatternFill("solid", fgColor=GREY)

# --- sheet 2: coverage ---------------------------------------------------
cv = wb.create_sheet("Capability Coverage")
cv.append(["Aberdeen capability", "Quals", "With quantified outcome", "Sellable with proof?"])
for k, v in cap_cov.most_common():
    o = cap_out.get(k, 0)
    cv.append([k, v, o, "STRONG" if o >= 2 else "MODERATE" if o == 1 else "THIN - capability only"])
cv.append([])
cv.append(["Industry", "Quals", "", ""])
for k, v in ind_cov.most_common():
    cv.append([k, v, "", ""])
for c in cv[1]:
    c.font = Font(bold=True, color="FFFFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=NAVY)
for i, w in enumerate([46, 10, 24, 26], 1):
    cv.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print("wrote: %s" % os.path.basename(OUT))
print("\nquals (excl. 2 non-quals): %d   with quantified outcome: %d (%.0f%%)"
      % (stats["quals"], stats["with_outcome"], 100.0 * stats["with_outcome"] / stats["quals"]))
print("\nCAPABILITY COVERAGE")
for k, v in cap_cov.most_common():
    print("   %-46s %2d quals, %d with outcome" % (k, v, cap_out.get(k, 0)))
print("\nINDUSTRY COVERAGE")
for k, v in ind_cov.most_common():
    print("   %-40s %2d" % (k, v))
